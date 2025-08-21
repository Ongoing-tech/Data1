from flask import Flask, render_template, request, jsonify, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
from werkzeug.utils import secure_filename
from config import Config
import os
import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from sqlalchemy import or_, and_, func
import re

app = Flask(__name__)
app.config.from_object(Config)

# 初始化数据库
db = SQLAlchemy(app)
CORS(app)

# 确保上传目录存在
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# 数据模型
class InventoryData(db.Model):
    __tablename__ = 'inventory_data'
    
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    date = db.Column(db.Date, nullable=False, index=True)
    sku = db.Column(db.String(50), nullable=False, index=True)
    product_name = db.Column(db.String(200), nullable=True)
    inbound_quantity = db.Column(db.Integer, nullable=True, default=0)
    outbound_quantity = db.Column(db.Integer, nullable=True, default=0)
    inventory_balance = db.Column(db.Integer, nullable=True, default=0)
    supplier = db.Column(db.String(100), nullable=True, index=True)
    operator = db.Column(db.String(50), nullable=True)
    remarks = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=db.func.now())
    updated_at = db.Column(db.DateTime, default=db.func.now(), onupdate=db.func.now())
    
    __table_args__ = (
        db.UniqueConstraint('date', 'sku', name='uq_date_sku'),
    )
    
    def to_dict(self):
        return {
            'id': self.id,
            'date': self.date.strftime('%Y-%m-%d') if self.date else None,
            'sku': self.sku,
            'product_name': self.product_name,
            'inbound_quantity': self.inbound_quantity,
            'outbound_quantity': self.outbound_quantity,
            'inventory_balance': self.inventory_balance,
            'supplier': self.supplier,
            'operator': self.operator,
            'remarks': self.remarks,
            'created_at': self.created_at.strftime('%Y-%m-%d %H:%M:%S') if self.created_at else None,
            'updated_at': self.updated_at.strftime('%Y-%m-%d %H:%M:%S') if self.updated_at else None
        }

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def parse_date_flexible(date_str):
    """灵活解析日期格式，支持多种格式"""
    date_formats = [
        '%Y-%m-%d',    # 2025-08-20
        '%Y/%m/%d',    # 2025/08/20
        '%Y.%m.%d',    # 2025.08.20
        '%Y%m%d',      # 20250820
        '%d-%m-%Y',    # 20-08-2025
        '%d/%m/%Y',    # 20/08/2025
        '%d.%m.%Y',    # 20.08.2025
        '%m-%d-%Y',    # 08-20-2025
        '%m/%d/%Y',    # 08/20-2025
        '%m.%d.%Y',    # 08.20.2025
    ]
    
    # 先尝试标准格式
    for fmt in date_formats:
        try:
            return datetime.datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    
    # 如果标准格式失败，尝试处理单数字月份和日期
    # 处理 YYYY/M/D 格式
    if re.match(r'^\d{4}/\d{1,2}/\d{1,2}$', date_str):
        parts = date_str.split('/')
        if len(parts) == 3:
            try:
                year = int(parts[0])
                month = int(parts[1])
                day = int(parts[2])
                return datetime.date(year, month, day)
            except ValueError:
                pass
    
    # 处理 YYYY-M-D 格式
    if re.match(r'^\d{4}-\d{1,2}-\d{1,2}$', date_str):
        parts = date_str.split('-')
        if len(parts) == 3:
            try:
                year = int(parts[0])
                month = int(parts[1])
                day = int(parts[2])
                return datetime.date(year, month, day)
            except ValueError:
                pass
    
    # 处理 YYYY.M.D 格式
    if re.match(r'^\d{4}\.\d{1,2}\.\d{1,2}$', date_str):
        parts = date_str.split('.')
        if len(parts) == 3:
            try:
                year = int(parts[0])
                month = int(parts[1])
                day = int(parts[2])
                return datetime.date(year, month, day)
            except ValueError:
                pass
    
    # 处理 D-M-YYYY 格式
    if re.match(r'^\d{1,2}-\d{1,2}-\d{4}$', date_str):
        parts = date_str.split('-')
        if len(parts) == 3:
            try:
                day = int(parts[0])
                month = int(parts[1])
                year = int(parts[2])
                return datetime.date(year, month, day)
            except ValueError:
                pass
    
    # 处理 D/M/YYYY 格式
    if re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', date_str):
        parts = date_str.split('/')
        if len(parts) == 3:
            try:
                day = int(parts[0])
                month = int(parts[1])
                year = int(parts[2])
                return datetime.date(year, month, day)
            except ValueError:
                pass
    
    # 处理 D.M.YYYY 格式
    if re.match(r'^\d{1,2}\.\d{1,2}\.\d{4}$', date_str):
        parts = date_str.split('.')
        if len(parts) == 3:
            try:
                day = int(parts[0])
                month = int(parts[1])
                year = int(parts[2])
                return datetime.date(year, month, day)
            except ValueError:
                pass
    
    return None

def validate_excel_data(data):
    """验证Excel数据格式"""
    errors = []
    
    for row_idx, row in enumerate(data, start=2):  # 从第2行开始（第1行是标题）
        if len(row) < 2:  # 至少需要date和sku
            errors.append(f"第{row_idx}行：数据不完整，至少需要日期和SKU")
            continue
            
        # 验证日期格式
        date_str = str(row[0]).strip()
        date = parse_date_flexible(date_str)
        if not date:
            errors.append(f"第{row_idx}行：日期格式错误，支持的格式包括：YYYY-MM-DD、YYYY/MM/DD、YYYY.MM.DD、YYYYMMDD、DD-MM-YYYY等")
            
        # 验证SKU
        sku = str(row[1]).strip()
        if not sku:
            errors.append(f"第{row_idx}行：SKU不能为空")
            
        # 验证数量字段（如果存在）
        if len(row) > 3:  # inbound_quantity
            try:
                if row[3]:
                    int(row[3])
            except (ValueError, TypeError):
                errors.append(f"第{row_idx}行：入库数量必须为整数")
                
        if len(row) > 4:  # outbound_quantity
            try:
                if row[4]:
                    int(row[4])
            except (ValueError, TypeError):
                errors.append(f"第{row_idx}行：出库数量必须为整数")
                
        if len(row) > 5:  # inventory_balance
            try:
                if row[5]:
                    int(row[5])
            except (ValueError, TypeError):
                errors.append(f"第{row_idx}行：库存余额必须为整数")
    
    return errors

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload')
def upload():
    return render_template('upload.html')

@app.route('/api/upload', methods=['POST'])
def upload_file():
    try:
        if 'file' not in request.files:
            return jsonify({'code': 400, 'message': '没有选择文件'}), 400
            
        file = request.files['file']
        if file.filename == '':
            return jsonify({'code': 400, 'message': '没有选择文件'}), 400
            
        if not allowed_file(file.filename):
            return jsonify({'code': 400, 'message': '文件格式不支持，只支持.xlsx格式'}), 400
            
        # 保存文件
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # 读取Excel文件
        try:
            workbook = openpyxl.load_workbook(filepath, data_only=True)
            sheet = workbook.active
            
            # 读取数据（跳过标题行）
            data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0]:  # 如果第一列（日期）有值
                    data.append(row)
            
            # 验证数据格式
            validation_errors = validate_excel_data(data)
            if validation_errors:
                return jsonify({
                    'code': 400,
                    'message': '数据格式验证失败',
                    'errors': validation_errors
                }), 400
            
            # 检查重复数据
            duplicate_items = []
            for row in data:
                if len(row) >= 2:
                    date_str = str(row[0]).strip()
                    sku = str(row[1]).strip()
                    
                    date = parse_date_flexible(date_str)
                    if date:
                        existing = InventoryData.query.filter_by(date=date, sku=sku).first()
                        if existing:
                            duplicate_items.append(f"SKU: {sku}, 日期: {date_str}")
            
            if duplicate_items:
                return jsonify({
                    'code': 400,
                    'message': f'导入失败：发现重复数据（{"; ".join(duplicate_items)}）'
                }), 400
            
            # 导入数据
            success_count = 0
            for row in data:
                if len(row) >= 2:
                    try:
                        date_str = str(row[0]).strip()
                        sku = str(row[1]).strip()
                        product_name = str(row[2]).strip() if len(row) > 2 and row[2] else None
                        inbound_quantity = int(row[3]) if len(row) > 3 and row[3] is not None else 0
                        outbound_quantity = int(row[4]) if len(row) > 4 and row[4] is not None else 0
                        inventory_balance = int(row[5]) if len(row) > 5 and row[5] is not None else 0
                        supplier = str(row[6]).strip() if len(row) > 6 and row[6] else None
                        operator = str(row[7]).strip() if len(row) > 7 and row[7] else None
                        remarks = str(row[8]).strip() if len(row) > 8 and row[8] else None
                        
                        date = parse_date_flexible(date_str)
                        if not date:
                            continue  # 跳过日期解析失败的行
                        
                        # 创建新记录
                        new_record = InventoryData(
                            date=date,
                            sku=sku,
                            product_name=product_name,
                            inbound_quantity=inbound_quantity,
                            outbound_quantity=outbound_quantity,
                            inventory_balance=inventory_balance,
                            supplier=supplier,
                            operator=operator,
                            remarks=remarks
                        )
                        
                        db.session.add(new_record)
                        success_count += 1
                        
                    except (ValueError, TypeError) as e:
                        continue
            
            db.session.commit()
            
            # 删除临时文件
            os.remove(filepath)
            
            return jsonify({
                'code': 200,
                'message': f'导入成功，共导入{success_count}条数据。'
            })
            
        except Exception as e:
            return jsonify({'code': 500, 'message': f'Excel文件读取失败：{str(e)}'}), 500
            
    except Exception as e:
        return jsonify({'code': 500, 'message': f'服务器错误：{str(e)}'}), 500

@app.route('/api/data', methods=['GET'])
def get_data():
    try:
        # 获取查询参数
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        sku = request.args.get('sku')
        product_name_like = request.args.get('product_name_like')
        supplier = request.args.get('supplier')
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 20))
        
        # 构建查询条件
        query = InventoryData.query
        
        if start_date:
            try:
                start_date_obj = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()
                query = query.filter(InventoryData.date >= start_date_obj)
            except ValueError:
                pass
                
        if end_date:
            try:
                end_date_obj = datetime.datetime.strptime(end_date, '%Y-%m-%d').date()
                query = query.filter(InventoryData.date <= end_date_obj)
            except ValueError:
                pass
                
        if sku:
            query = query.filter(InventoryData.sku == sku)
            
        if product_name_like:
            query = query.filter(InventoryData.product_name.like(f'%{product_name_like}%'))
            
        if supplier:
            query = query.filter(InventoryData.supplier == supplier)
        
        # 按日期降序排序
        query = query.order_by(InventoryData.date.desc())
        
        # 分页查询
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        
        # 构建响应数据
        result = {
            'code': 200,
            'data': {
                'items': [item.to_dict() for item in pagination.items],
                'total': pagination.total,
                'page': page,
                'per_page': per_page,
                'pages': pagination.pages
            }
        }
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'code': 500, 'message': f'查询失败：{str(e)}'}), 500

@app.route('/api/suppliers', methods=['GET'])
def get_suppliers():
    try:
        # 获取所有不重复的供应商列表
        suppliers = db.session.query(InventoryData.supplier)\
            .filter(InventoryData.supplier.isnot(None))\
            .filter(InventoryData.supplier != '')\
            .distinct()\
            .order_by(InventoryData.supplier)\
            .all()
        
        supplier_list = [supplier[0] for supplier in suppliers]
        
        return jsonify({
            'code': 200,
            'data': supplier_list
        })
        
    except Exception as e:
        return jsonify({'code': 500, 'message': f'获取供应商列表失败：{str(e)}'}), 500

@app.route('/api/chart-data', methods=['GET'])
def get_chart_data():
    try:
        # 获取查询参数
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        sku = request.args.get('sku')
        supplier = request.args.get('supplier')
        
        # 构建查询条件
        query = InventoryData.query
        
        if start_date:
            try:
                start_date_obj = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()
                query = query.filter(InventoryData.date >= start_date_obj)
            except ValueError:
                pass
                
        if end_date:
            try:
                end_date_obj = datetime.datetime.strptime(end_date, '%Y-%m-%d').date()
                query = query.filter(InventoryData.date <= end_date_obj)
            except ValueError:
                pass
                
        if sku:
            query = query.filter(InventoryData.sku == sku)
            
        if supplier:
            query = query.filter(InventoryData.supplier == supplier)
        
        # 按日期分组，计算每日库存余额总和
        chart_data = query.with_entities(
            InventoryData.date,
            func.sum(InventoryData.inventory_balance).label('total_balance')
        ).group_by(InventoryData.date).order_by(InventoryData.date).all()
        
        # 格式化数据
        dates = [item.date.strftime('%Y-%m-%d') for item in chart_data]
        balances = [float(item.total_balance) if item.total_balance else 0 for item in chart_data]
        
        return jsonify({
            'code': 200,
            'data': {
                'dates': dates,
                'balances': balances
            }
        })
        
    except Exception as e:
        return jsonify({'code': 500, 'message': f'获取图表数据失败：{str(e)}'}), 500

if __name__ == '__main__':
    app.run(debug=app.config['DEBUG'], host='0.0.0.0', port=5000)